import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side,PatternFill
import os.path
from pathlib import Path

import xml.etree.ElementTree as ET



from .fonction import afficheerreur
from .constante import *


class Tableur:

    def __init__(self):
        # classeur
        self.wbook = None
        # feuille
        self.wsfeuille = None


    def initclasseur(self):
        fichier_excel = Path(os.path.dirname(__file__) + "/transaction.xlsx")

        tree = ET.parse(PATH_REP + "\XML\Attributs.xml")
        root = tree.getroot()
        list_champs = []
        premiere_lettre = "D"
        for champs in root.findall("champs"):
            list_champs.append(champs.get("id"))

        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        if not fichier_excel.exists():
            self.wbook = openpyxl.Workbook()
            self.wsfeuille = self.wbook.active
            self.wsfeuille.title = "Assistant DFCI"
            self.wsfeuille["A1"] = "Date"
            self.wsfeuille["B1"] = "N° de transaction"
            self.wsfeuille["C1"] = "CLEABS"
            self.wsfeuille["D1"] = "Champs modifiés"
            self.wsfeuille.column_dimensions["A"].width = 25
            self.wsfeuille.column_dimensions["B"].width = 28
            self.wsfeuille.column_dimensions["C"].width = 28
            self.wsfeuille.column_dimensions["D"].width = 100

            # for champs in list_champs:
            #     numero_cellule = f"{premiere_lettre}1"
            #     self.wsfeuille[numero_cellule] = champs
            #     # definition de la taille de la colonne
            #     self.wsfeuille.column_dimensions[premiere_lettre].width = 40
            #     self.wsfeuille[numero_cellule].font = bold_font
            #     self.wsfeuille[numero_cellule].alignment = center_alignment
            #     # recup du code ascii , incrementation et retransformation en lettre.
            #     premiere_lettre= chr(ord(premiere_lettre) + 1)

            # Formater l'aspect des cellules
            thick_border = Border(left=Side(style="thick"),
                                 right=Side(style="thick"),
                                 top=Side(style="thick"),
                                 bottom=Side(style="thick"))

            # Appliquer les bordures à toutes les cellules contenant des données
            for row in self.wsfeuille.iter_rows(min_row=1, max_row=self.wsfeuille.max_row, min_col=1, max_col=self.wsfeuille.max_column):
                for cell in row:
                    cell.border = thick_border

            # Formater la colonne A (colonne 1) pour qu'Excel reconnaisse bien les dates
            for row in self.wsfeuille.iter_rows(min_row=1, max_row=self.wsfeuille.max_row, min_col=1, max_col=1):
                for cell in row:
                    # Définir le format de la cellule pour afficher les dates au format 'YYYY-MM-DD'
                    cell.number_format = 'YYYY-MM-DD'

        else:
            self.wbook = openpyxl.load_workbook(fichier_excel)
            self.wsfeuille = self.wbook["Assistant DFCI"]


    def log_is_open(self):
        # on test si le fichier est deja ouvert
        try:
            # Tente d'ouvrir le fichier en mode écriture
            with open(os.path.dirname(__file__) + "/transaction.xlsx", 'r+'):
                # le fichier n'est pas ouvert
                return True

        except PermissionError:
            afficheerreur(
                "La modification n'a pas été appliquée<br>Le fichier de <b>log</b> : <mark style='background-color: "
                "lightgreen;'>transaction.xlsx</mark> est ouvert<p><br>Veuillez le fermer pour pouvoir ecrire de "
                "nouvelles données")

            return False  # Une erreur signifie que le fichier est déjà ouvert
        except FileNotFoundError:
            # le teste d'existance ce fait dans initclasseur
            return True


    def adddonnees(self,listdonnees):

        self.initclasseur()

        cell_color = PatternFill(start_color="ff978d", end_color="ff978d", fill_type="solid")
        couple_cell = [["D2", "E2"], ["F2", "G2"]]

        for ligne in listdonnees:
            # print(ligne)
            # self.wsfeuille.append(ligne)
            # inserer une ligne vide a la position 2 (premiere ligne apres les titres)
            self.wsfeuille.insert_rows(2)
            # self.wsfeuille.cell(row=2, column=4, value="test\ntest1")
            # ajout des données dans la ligne crée
            for col, valeur in enumerate(ligne, start=1):
                self.wsfeuille.cell(row=2, column=col, value=valeur)

            # colorier les celulles dont les attributs ont changés
            # for couple in couple_cell:
            #     cellule1 = self.wsfeuille[couple[0]]
            #     cellule2 = self.wsfeuille[couple[1]]
            #     if cellule1.value != cellule2.value:
            #         # afficheerreur("nature differente , on colorie")
            #         cellule1.fill = cell_color
            #         cellule2.fill = cell_color

        for row in self.wsfeuille.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
        return True


    def sauvegarder(self):
        self.wbook.save(os.path.dirname(__file__) + "/transaction.xlsx")


